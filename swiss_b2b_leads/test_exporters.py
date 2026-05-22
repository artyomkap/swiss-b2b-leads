import sys
sys.path.insert(0, '.')

from models import Lead
from exporters.csv_exporter import export as csv_export
from exporters.excel_exporter import export as excel_export

leads = [
    Lead(company_name='Müller Garage AG', city='Zurich', phone='+41441234567', email='info@mueller.ch', website='https://mueller.ch', source='search.ch', quality_score=85),
    Lead(company_name='Geneva Auto SA', city='Geneva', phone='+41221234567', website='https://genevaauto.ch', source='google_search', quality_score=55),
]
source_stats = [
    {'source': 'search.ch', 'records_collected': 5, 'records_after_deduplication': 4, 'emails_found': 3, 'phones_found': 4, 'websites_found': 4, 'average_quality_score': 72.5, 'notes': ''},
]

csv_export(leads, 'output/test_leads.csv')
excel_export(leads, 'output/test_leads.xlsx', source_stats)

import os, csv
with open('output/test_leads.csv', encoding='utf-8-sig') as f:
    rows = list(csv.DictReader(f))
print(f'CSV: {len(rows)} rows, fields: {list(rows[0].keys())[:5]}...')

import openpyxl
wb = openpyxl.load_workbook('output/test_leads.xlsx')
print(f'Excel sheets: {wb.sheetnames}')
ws = wb["leads"]
print(f'Excel leads rows: {ws.max_row - 1}')
